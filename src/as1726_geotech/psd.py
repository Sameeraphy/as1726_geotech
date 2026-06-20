"""Particle Size Distribution (PSD) utilities.

This module provides utilities to compute percent passing, D10/D30/D60
statistics and coefficients of uniformity/curvature from either
retained masses or percent-passing data.
"""
from __future__ import annotations

from typing import Iterable, List, Optional, Tuple, Dict
import math


def percent_passing_from_retained(sieve_sizes: Iterable[float], retained: Iterable[float]) -> Tuple[List[float], List[float]]:
    """Compute percent passing given sieve sizes and retained masses.

    Args:
        sieve_sizes: iterable of sieve sizes (mm). Can be any order; will be
            sorted descending (largest to smallest) internally.
        retained: iterable of masses retained on each sieve (same order as
            `sieve_sizes`).

    Returns:
        (sizes_sorted, percent_passing): both lists sorted from largest to
        smallest sieve size. Percent passing is in percent (0-100).
    """
    sizes = list(sieve_sizes)
    ret = list(retained)
    if len(sizes) != len(ret):
        raise ValueError("sieve_sizes and retained must have the same length")

    # Pair and sort by size descending
    paired = sorted(zip(sizes, ret), key=lambda x: x[0], reverse=True)
    sizes_sorted = [p[0] for p in paired]
    retained_sorted = [p[1] for p in paired]

    total = sum(retained_sorted)
    if total <= 0:
        raise ValueError("Total mass must be positive")

    percent_passing = []
    cumulative_retained = 0.0
    for r in retained_sorted:
        # percent finer than this sieve = (total - cumulative_retained_up_to_prev) / total *100
        percent = (total - cumulative_retained) / total * 100.0
        percent_passing.append(percent)
        cumulative_retained += r
    return sizes_sorted, percent_passing


def interpolate_d(sizes: Iterable[float], percent_passing: Iterable[float], target_percent: float) -> Optional[float]:
    """Interpolate particle size for a given percent passing.

    Uses linear interpolation on log10(size) vs percent passing.

    Args:
        sizes: sieve sizes sorted descending (largest to smallest).
        percent_passing: percent passing corresponding to each sieve.
        target_percent: target percent passing (0-100).

    Returns:
        Particle size (same units as `sizes`) where percent passing equals
        `target_percent`, or `None` if target is outside the data range.
    """
    s = list(sizes)
    p = list(percent_passing)
    if len(s) != len(p):
        raise ValueError("sizes and percent_passing must have same length")

    # Ensure lists are sorted descending by size
    paired = sorted(zip(s, p), key=lambda x: x[0], reverse=True)
    s = [x[0] for x in paired]
    p = [x[1] for x in paired]

    # If target equals one of the tabulated values, return exact size
    for si, pi in zip(s, p):
        if math.isclose(pi, target_percent, rel_tol=1e-9, abs_tol=1e-12):
            return float(si)

    # find bracket where p[i] >= target >= p[i+1] (since p increases as size decreases)
    for i in range(len(s) - 1):
        p1, p2 = p[i], p[i + 1]
        if (p1 >= target_percent >= p2) or (p2 >= target_percent >= p1):
            s1, s2 = s[i], s[i + 1]
            # avoid divide by zero
            if math.isclose(p1, p2):
                # percent plateau — return geometric mean
                return math.sqrt(s1 * s2)
            # interpolate on log(size)
            log1, log2 = math.log10(s1), math.log10(s2)
            t = (target_percent - p1) / (p2 - p1)
            logx = log1 + t * (log2 - log1)
            return 10 ** logx

    return None


def compute_gradation(sieve_sizes: Iterable[float], retained: Optional[Iterable[float]] = None, percent_passing: Optional[Iterable[float]] = None) -> dict:
    """Compute gradation summary including D10, D30, D60, Cu and Cc.

    Provide either `retained` masses or `percent_passing` values.

    Returns:
        dict with keys: `sizes`, `percent_passing`, `D10`, `D30`, `D60`, `Cu`, `Cc`.
    """
    if retained is None and percent_passing is None:
        raise ValueError("Provide either retained masses or percent_passing")

    if retained is not None:
        sizes, p = percent_passing_from_retained(sieve_sizes, retained)
    else:
        # use provided percent_passing — sort inputs
        paired = sorted(zip(sieve_sizes, percent_passing), key=lambda x: x[0], reverse=True)
        sizes = [x[0] for x in paired]
        p = [x[1] for x in paired]

    D10 = interpolate_d(sizes, p, 10.0)
    D30 = interpolate_d(sizes, p, 30.0)
    D60 = interpolate_d(sizes, p, 60.0)

    Cu = None
    Cc = None
    if D10 and D60 and D10 > 0:
        Cu = D60 / D10
        Cc = (D30 * D30) / (D10 * D60) if (D30 is not None) else None

    return {
        "sizes": sizes,
        "percent_passing": p,
        "D10": D10,
        "D30": D30,
        "D60": D60,
        "Cu": Cu,
        "Cc": Cc,
    }


def plot_gradation(sizes: Iterable[float], percent_passing: Iterable[float], title: str = "Particle Size Distribution", show: bool = True):
    """Plot PSD curve (semilog: log size vs linear percent).

    Args:
        sizes: sieve sizes (mm).
        percent_passing: percent passing values (0-100).
        title: plot title.
        show: if True, display plot; else return figure.

    Returns:
        matplotlib figure if show=False, else None.
    """
    try:
        import matplotlib.pyplot as plt
    except ImportError:
        raise ImportError("matplotlib is required for plotting. Install with: pip install matplotlib")

    fig, ax = plt.subplots(figsize=(10, 6))
    s = list(sizes)
    p = list(percent_passing)
    # Sort by size
    paired = sorted(zip(s, p), key=lambda x: x[0], reverse=True)
    s = [x[0] for x in paired]
    p = [x[1] for x in paired]

    ax.semilogx(s, p, 'b-o', linewidth=2, markersize=6, label="PSD")
    ax.set_xlabel("Particle Size (mm)", fontsize=12)
    ax.set_ylabel("Percent Passing (%)", fontsize=12)
    ax.set_title(title, fontsize=14, fontweight="bold")
    ax.grid(True, which="both", alpha=0.3)
    ax.legend()
    ax.set_ylim([0, 100])

    if show:
        plt.show()
        return None
    else:
        return fig


def load_from_excel(file_path: str, sieve_col: str = "Sieve (mm)", retained_col: str = "Retained (g)") -> Dict[str, List[float]]:
    """Load PSD data from Excel file.

    Expected Excel format:
    - Column headers: "Sieve (mm)" and "Retained (g)" (customizable)
    - Rows: sieve size and corresponding retained mass

    Args:
        file_path: path to Excel file (.xlsx).
        sieve_col: name of sieve size column.
        retained_col: name of retained mass column.

    Returns:
        dict with keys "sieves" and "retained" containing lists of values.

    Example:
        >>> data = load_from_excel("sample.xlsx")
        >>> sizes, percent = percent_passing_from_retained(data["sieves"], data["retained"])
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for Excel support. Install with: pip install openpyxl")

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Find column indices
    sieve_idx = None
    retained_idx = None
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value and str(cell.value).strip().lower() == sieve_col.lower():
            sieve_idx = col_idx
        if cell.value and str(cell.value).strip().lower() == retained_col.lower():
            retained_idx = col_idx

    if sieve_idx is None or retained_idx is None:
        raise ValueError(f"Columns '{sieve_col}' and/or '{retained_col}' not found in Excel file")

    sieves = []
    retained = []

    # Read data rows (skip header)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if row[sieve_idx - 1] is not None and row[retained_idx - 1] is not None:
            try:
                sieves.append(float(row[sieve_idx - 1]))
                retained.append(float(row[retained_idx - 1]))
            except (ValueError, TypeError):
                continue

    if not sieves:
        raise ValueError("No valid data found in Excel file")

    return {"sieves": sieves, "retained": retained}


# Backwards-compatible simple function name
particle_size_distribution = compute_gradation
